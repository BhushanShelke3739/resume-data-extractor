import streamlit as st
import os
import re
import PyPDF2
import openpyxl
from docx import Document
from docx2pdf import convert
from tika import parser


def extract_data(file):
    """
    Extracts data (emails, phone numbers, text content) from a file.
    """
    text = ""
    if file.name.endswith(".pdf"):
        pdf_reader = PyPDF2.PdfReader(file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()

            

    elif file.name.endswith(".docx"):
        doc = Document(file)
        text = "\n".join([para.text for para in doc.paragraphs])

    elif file.name.endswith(".doc"):
        parsed = parser.from_buffer(file.read()) 
        text = parsed["content"]

    email_ids = re.findall(r"[\w\.-]+@[\w\.-]+", text)
    contact_numbers = re.findall(r"\b(?:\+\d{1,2}\s?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}\b", text)
    return text, email_ids, contact_numbers


def create_excel(data):
  """
  Creates an Excel workbook with extracted data.
  """
  workbook = openpyxl.Workbook()
  worksheet = workbook.active

  worksheet["A1"] = "File Name"
  worksheet["B1"] = "Email IDs"
  worksheet["C1"] = "Contact Numbers"
  worksheet["D1"] = "Text Content"

  row = 2
  for filename, text, email_ids, contact_numbers in data:
    worksheet.cell(row=row, column=1, value=filename)
    worksheet.cell(row=row, column=2, value=", ".join(email_ids))
    worksheet.cell(row=row, column=3, value=", ".join(contact_numbers))
    worksheet.cell(row=row, column=4, value=text)
    row += 1

  return workbook

st.title("Resume Data Extractor")
uploaded_files = st.file_uploader("Upload Resumes", type=["pdf", "docx", "doc"], accept_multiple_files=True)

if uploaded_files:
  data = []
  for file in uploaded_files:
    text, email_ids, contact_numbers = extract_data(file)
    data.append((file.name, text, email_ids, contact_numbers))

  if data:
    workbook = create_excel(data)
    download_excel = st.button("Download Extracted Data (.xls)")

    if download_excel:
      with st.spinner("Downloading..."):
        filename = "extracted_data.xls"
        workbook.save(filename)
        st.success(f"Data downloaded as {filename}")
        with open(filename, "rb") as f:
          st.download_button("Download", f, file_name=filename)
        os.remove(filename)  
  else:
    st.info("No data found in uploaded files.")
